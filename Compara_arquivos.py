## propósito: verificar na planilha, quais maquinas não estão sendo monitoradas e são de PRD.
## procedimento: usar como base a planilha dos Nao_monitorados, localizar na planilha de ativos e verificar na coluna "Cluster" se é PRD. Por fim listar.

import time
import openpyxl

## Acessando os NaoMonitorados
inicio = time.time()
path = r"Aqui entra o caminho do arquivo Excel.xlsx"
wb_obj = openpyxl.load_workbook(path, data_only = True) 
sheet_obj = wb_obj.active 
max_col = sheet_obj.max_column ## total de colunas
max_row = sheet_obj.max_row  ## total de linhas

## Acessando o LevantamentoAtivos

path2 = r"Aqui entra o segundo arquivo.xlsx"
wb_obj2 = openpyxl.load_workbook(path2, data_only = True)
sheet_obj2 = wb_obj2.active
max_col2 = sheet_obj2.max_column ## total de colunas
max_row2= sheet_obj2.max_row  ## total de linhas

# declarando algumas variaveis
NaoEncontradas = []
cont=0
cont2=0
contador=0
encontrada = False
#print(max_row,"\n")
#print(max_row2,"\n")
print("Máquinas de PRD que estão ligadas e não estão sendo monitoradas!\n")

# Fazendo a checagem!
for i in range(2, max_row + 1): 
    cell_obj = sheet_obj.cell(row = i, column = 2)
    for j in range(1,max_row2+1):
        cell_obj2 = sheet_obj2.cell(row = j, column = 1)
        
        on_off = sheet_obj2.cell(row = j, column = 2)
        cluster = sheet_obj2.cell(row = j, column = 62)
        OPRO = sheet_obj2.cell(row = j, column =93)        
        
        if cell_obj.value == cell_obj2.value and on_off.value == 'poweredOn' and cluster.value == 'PRD':
            #print(on_off.value," | ",cluster.value," | ",OPRO.value,"\n")
            print(on_off.value,",",cell_obj2.value,", Encontrada na linha:",j+1,",",cluster.value,",",OPRO.value)
            contador+=1
            #if on_off.value == 'poweredOn' and cluster.value == 'PRD':
            #    cont+=1
            #    print(sheet_obj2.cell(row = j, column = 1).value,
            #          ",",sheet_obj2.cell(row = j, column = 2).value,
            #          ",",sheet_obj2.cell(row = j, column = 62).value,
            #          ",",sheet_obj2.cell(row = j, column = 93).value)
            break  


print("\n\nQuantidade de maquinas testadas:",contador)
print("\n\n\n\n\n ")


for i in range(2, max_row + 1): 
    cell_obj = sheet_obj.cell(row = i, column = 2)
    for j in range(1,max_row2+1):
        cell_obj2 = sheet_obj2.cell(row = j, column = 1)
        
        on_off = sheet_obj2.cell(row = j, column = 2)
        cluster = sheet_obj2.cell(row = j, column = 62)
        OPRO = sheet_obj2.cell(row = j, column =93)
        
        #and on_off.value == 'poweredOn' ## inserir este and na condicional abaixo para pegar apenas as ligadas!
        
        if cell_obj.value == cell_obj2.value and on_off.value == 'poweredOn':
            encontrada=True
            cont+=1
            print(on_off.value,",",cell_obj2.value,", Encontrada na linha:",j+1,",",cluster.value,",",OPRO.value)
        if encontrada == False and j == max_row2:
            cont2+=1
            string = str(cell_obj.value)+",Encontrada na linha:"+str(i)
            NaoEncontradas.append(string)
            break
    encontrada=False
    
print("\n\n")

for i in range(0,len(NaoEncontradas)):
    print("Sem Status, ",NaoEncontradas[i])
## Imprimindo o rodapé

print("\nTotal:",cont,"máquinas Encontradas")
print("Total:",cont2,"máquinas Nao Encontradas\n")
print("Maquinas a testar = {};\nMaquinas testadas = {};\nUm total de {} maquinas deixaram de ser testadas."
      .format(max_row-1,cont+cont2,(max_row-1)-(cont+cont2)))
fim = time.time()
print("A execução durou %.2f" % (fim - inicio),"segundos")
